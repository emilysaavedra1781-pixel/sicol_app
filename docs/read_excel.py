"""Script para leer el Excel Casos_Prueba.xlsx y generar inventario en JSON."""
import openpyxl
import json
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

wb = openpyxl.load_workbook('docs/Casos_Prueba.xlsx', data_only=True)

inventory = {}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows_data = []
    for row in ws.iter_rows(values_only=True):
        if any(cell is not None for cell in row):
            # Convertir a string seguro
            safe_row = []
            for cell in row:
                if cell is None:
                    safe_row.append(None)
                else:
                    safe_row.append(str(cell))
            rows_data.append(safe_row)
    inventory[sheet_name] = rows_data

# Guardar como JSON
with open('docs/casos_prueba_inventory.json', 'w', encoding='utf-8') as f:
    json.dump(inventory, f, ensure_ascii=False, indent=2)

print(f"Total de pestañas: {len(inventory)}")
for sheet, rows in inventory.items():
    print(f"  {sheet}: {len(rows)} filas (incl. encabezados)")
print("Guardado en docs/casos_prueba_inventory.json")
